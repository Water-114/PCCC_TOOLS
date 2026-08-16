"""Xuất data/ai-giang-day-densucco.json từ "AI_GIANG_DAY/NGUON_SACH_PCCC (6).xlsx"
(sheet NOI_DUNG_BAN_VE, lọc he_thong chứa CHÍNH XÁC token B12, loại trừ các
dòng ĐỒNG THỜI chứa B3/B5/B6 vì đã thuộc mục "3. Chữa cháy nước" rồi — tránh
trùng lặp nội dung ở 2 mục) — mục "6. Bản vẽ thiết kế đèn CSSC, đèn Exit, sơ
đồ thoát nạn, bình chữa cháy" trong tab "AI giảng dạy" → tab con 2 (phần
bình chữa cháy — nguồn này CHƯA có dữ liệu đèn sự cố/đèn Exit/thoát nạn).
Theo đúng khuôn prepare_ai_giang_day_ccnuoc.py: chạy 1 lần (hoặc lại khi
owner cập nhật file Excel/thư viện ảnh), KHÔNG chạy trong Flask lúc request,
dùng openpyxl hệ thống (KHÔNG có trong requirements.txt).

Lọc he_thong: tách theo dấu "/", so khớp CHÍNH XÁC token "B12" VÀ không có
token nào trong {B3, B5, B6} (loại trừ dòng TK-01 "Bảng thống kê tổng hợp
thiết bị", he_thong="B3/B5/B6/B12", đã thuộc mục ccnuoc). Bỏ qua dòng không
có ma_muc. Kỳ vọng: đúng 3 dòng (BC-01/02/03, "Mặt bằng bình chữa cháy").

Ảnh minh hoạ: xử lý y hệt logic ccnuoc (os.path.exists trên
AI_GIANG_DAY/THU_VIEN_ANH/{ten_file_anh}, giữ nguyên cấu trúc thư mục con
khi copy) dù dữ liệu hiện tại của mục này chắc chắn không có ảnh nào (cả 3
dòng ten_file_anh đều rỗng) — giữ cùng logic để nhất quán, không cần sửa gì
nếu owner bổ sung thêm dòng có ảnh sau này.

Chạy: python scripts/prepare_ai_giang_day_densucco.py
"""

import json
import shutil
import sys
import io
from pathlib import Path

from openpyxl import load_workbook

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SOURCE_XLSX = REPO_ROOT / "AI_GIANG_DAY" / "NGUON_SACH_PCCC (6).xlsx"
SHEET_NAME = "NOI_DUNG_BAN_VE"
IMG_SOURCE_DIR = REPO_ROOT / "AI_GIANG_DAY" / "THU_VIEN_ANH"
IMG_DEST_DIR = Path(__file__).resolve().parent.parent / "app" / "static" / "img" / "ai-giang-day" / "densucco"
DEST_JSON = Path(__file__).resolve().parent.parent / "app" / "static" / "data" / "ai-giang-day-densucco.json"

TARGET_TOKEN = "B12"
EXCLUDE_TOKENS = {"B3", "B5", "B6"}


def _s(v):
    """Chuẩn hoá 1 ô Excel: None hoặc chuỗi toàn khoảng trắng -> None, nếu không thì strip()."""
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def _he_thong_tokens(he_thong):
    if not he_thong:
        return set()
    return {t.strip() for t in str(he_thong).split("/") if t.strip()}


def main():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    assert header[:10] == (
        "ma_muc", "loai_ban_ve", "he_thong", "noi_dung", "yeu_cau_chi_tiet",
        "can_cu_phap_ly", "loi_thuong_gap", "ten_file_anh", "nguon_anh", "trang_thai",
    ), header

    out = []
    n_img_real = 0
    n_img_null = 0
    no_file_found = []
    for ma_muc, loai_ban_ve, he_thong, noi_dung, yeu_cau, can_cu, loi, ten_file_anh, nguon_anh, trang_thai in (
        r[:10] for r in data_rows
    ):
        ma_muc = _s(ma_muc)
        if not ma_muc:
            continue
        tokens = _he_thong_tokens(he_thong)
        if TARGET_TOKEN not in tokens or (tokens & EXCLUDE_TOKENS):
            continue

        raw_ten_file_anh = _s(ten_file_anh)
        resolved_ten_file_anh = None
        if raw_ten_file_anh:
            src_path = IMG_SOURCE_DIR / raw_ten_file_anh
            if src_path.is_file():
                dest_path = IMG_DEST_DIR / raw_ten_file_anh
                dest_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(src_path, dest_path)
                resolved_ten_file_anh = raw_ten_file_anh.replace("\\", "/")
                n_img_real += 1
            else:
                n_img_null += 1
                no_file_found.append((ma_muc, raw_ten_file_anh))
        else:
            n_img_null += 1

        out.append({
            "ma_muc": ma_muc,
            "loai_ban_ve": _s(loai_ban_ve),
            "he_thong": _s(he_thong),
            "noi_dung": _s(noi_dung),
            "yeu_cau_chi_tiet": _s(yeu_cau),
            "can_cu_phap_ly": _s(can_cu),
            "loi_thuong_gap": _s(loi),
            "ten_file_anh": resolved_ten_file_anh,
            "nguon_anh": _s(nguon_anh),
            "trang_thai": _s(trang_thai),
        })

    DEST_JSON.parent.mkdir(parents=True, exist_ok=True)
    DEST_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    groups = sorted({r["loai_ban_ve"] for r in out if r["loai_ban_ve"]})
    print(f"Da doc {len(data_rows)} dong tho tu {SHEET_NAME}")
    print(f"  - Loc con {len(out)} dong (he_thong chua dung token B12, loai tru B3/B5/B6) -> ghi vao {DEST_JSON}")
    print(f"  - So nhom loai_ban_ve phan biet: {len(groups)}")
    for g in groups:
        print(f"      - {g}")
    print(f"  - Anh thuc su copy duoc: {n_img_real}")
    print(f"  - ten_file_anh = null (rong hoac khong tim thay file): {n_img_null}")
    if no_file_found:
        print(f"  - Danh sach dong CO CHU o ten_file_anh nhung KHONG tim thay file that ({len(no_file_found)}):")
        for ma_muc, raw in no_file_found:
            print(f"      {ma_muc}: {raw!r}")


if __name__ == "__main__":
    main()
