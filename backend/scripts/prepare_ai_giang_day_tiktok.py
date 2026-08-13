"""Xuất data/ai-giang-day-tiktok.json từ AI_GIANG_DAY/DANH_SACH_LINK_TIKTOK.xlsx
(sheet BAI_GIANG_TIKTOK) — chạy 1 lần (hoặc lại khi owner cập nhật file Excel),
KHÔNG chạy trong Flask lúc request. Dùng openpyxl (KHÔNG có trong
requirements.txt — chỉ dùng cho script chuẩn bị dữ liệu 1 lần này, không phải
dependency runtime của app, giống cách python-docx dùng ở prepare_mdc_templates.py
nhưng openpyxl không cần thiết lúc runtime nên cố tình không thêm vào
requirements.txt).

Lọc bỏ:
- Dòng ví dụ mẫu (ghi_chu chứa "Ví dụ mẫu").
- Dòng tiktok_url rỗng (chưa dán link).

Trích video_id từ tiktok_url bằng regex ".../video/{id}" — nếu không khớp
định dạng (link rút gọn/dạng khác), video_id = null để frontend tự hiển thị
fallback "Xem trên TikTok" thay vì nhúng player.

Chạy: python scripts/prepare_ai_giang_day_tiktok.py
"""

import json
import re
import sys
import io
from pathlib import Path

from openpyxl import load_workbook

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SOURCE_XLSX = Path(__file__).resolve().parent.parent.parent / "AI_GIANG_DAY" / "DANH_SACH_LINK_TIKTOK.xlsx"
SHEET_NAME = "BAI_GIANG_TIKTOK"
DEST_JSON = Path(__file__).resolve().parent.parent / "app" / "static" / "data" / "ai-giang-day-tiktok.json"

VIDEO_ID_RE = re.compile(r"/video/(\d+)")


def _s(v):
    """Chuẩn hoá 1 ô Excel: None hoặc chuỗi toàn khoảng trắng -> None, nếu không thì strip()."""
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def main():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    assert header[:5] == ("stt", "tiktok_url", "tieu_de_de_xuat", "chuyen_de", "ghi_chu"), header

    out = []
    skipped_mau = 0
    skipped_empty_url = 0
    for stt, tiktok_url, tieu_de_de_xuat, chuyen_de, ghi_chu in (r[:5] for r in data_rows):
        ghi_chu = _s(ghi_chu)
        if ghi_chu and "Ví dụ mẫu" in ghi_chu:
            skipped_mau += 1
            continue
        tiktok_url = _s(tiktok_url)
        if not tiktok_url:
            skipped_empty_url += 1
            continue

        m = VIDEO_ID_RE.search(tiktok_url)
        video_id = m.group(1) if m else None

        tieu_de = _s(tieu_de_de_xuat) or ("Video PCCC #" + str(stt))

        out.append({
            "stt": stt,
            "tiktok_url": tiktok_url,
            "tieu_de": tieu_de,
            "chuyen_de": _s(chuyen_de),
            "ghi_chu": ghi_chu,
            "video_id": video_id,
        })

    DEST_JSON.parent.mkdir(parents=True, exist_ok=True)
    DEST_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    no_video_id = [r for r in out if r["video_id"] is None]
    print(f"Da doc {len(data_rows)} dong du lieu tu {SHEET_NAME}")
    print(f"  - Bo qua {skipped_mau} dong vi du mau")
    print(f"  - Bo qua {skipped_empty_url} dong tiktok_url rong")
    print(f"  - Con lai {len(out)} dong -> ghi vao {DEST_JSON}")
    print(f"  - So dong video_id=null (fallback 'Xem tren TikTok'): {len(no_video_id)}")
    for r in no_video_id:
        print(f"      stt={r['stt']} url={r['tiktok_url']}")

    distinct_cd = sorted({r["chuyen_de"] for r in out if r["chuyen_de"]})
    n_unclassified = sum(1 for r in out if not r["chuyen_de"])
    print(f"  - Chuyen de phan biet: {distinct_cd}")
    print(f"  - So dong chua phan loai (chuyen_de rong): {n_unclassified}")


if __name__ == "__main__":
    main()
